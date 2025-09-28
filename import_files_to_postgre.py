import numpy as np
import os
import pandas as pd
import psycopg2
from openpyxl import load_workbook
# Importar funciones generales
from utils_tools import (
    PROCESSED_DIR,
    ERRORS_DIR,
    print_log,
    safe_convert_to_float,
    move_file
)
# Importar funciones para DB
from database_utils import (
    DB_CONFIG,
    CAT_PAYMENT_TYPE,
    CAT_STORE,
    get_catalogs,
    get_id_payment_type,
    get_or_create_store,
    get_or_create_provider,
    create_product,
    insert_purchase,
    insert_operations,
    insert_price
)

MARGEN_GANANCIA = 0.30  # 30% de margen de ganancia
DESCUENTO_OFERTA = 0.15  # 15% de descuento en ofertas

# Habilitar/Dehabilitar LOGs
ENABLE_LOGS = True

# ==== DIRECTORIO DE ARCHIVOS ====
DATA_DIR = "data_files_ingestion"

# Crear directorios si no existen
os.makedirs(PROCESSED_DIR, exist_ok=True)
os.makedirs(ERRORS_DIR, exist_ok=True)

# Mapeo de URLs
PICTURE_URL = []

def extract_hyperlinks(wb, sheetName="Precios", columna="Preview"):
    """Versión modificada para trabajar con workbook ya abierto"""
    try:
        ws = wb[sheetName]
        headers = [str(cell.value).strip() if cell.value else "" 
                  for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        try:
            col_idx = headers.index(columna) + 1
        except ValueError:
            print_log(f"⚠️ Columna '{columna}' no encontrada")
            return [""] * (ws.max_row - 1)
        urls = []
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            cell = row[0]
            url = cell.hyperlink.target if cell.hyperlink else ""
            urls.append(url)
        print_log(f"Extraídos {len(urls)} URLs")
        return urls
    except Exception as e:
        print_log(f"❌ Error extrayendo hipervínculos: {str(e)}")
        return []

def verify_columns(df, required_columns, df_name=""):
    """Verifica que un DataFrame tenga las columnas requeridas."""
    missing = [col for col in required_columns if col not in df.columns]
    if missing:
        print_log(f"⚠️ Columnas faltantes en {df_name}: {missing}")
        return False
    return True

def procesar_purchase(dfPrchss, dfPrices):
    """Procesa el dataframe de purchase con manejo robusto de errores."""
    df = dfPrchss.copy()
    # Debug: Mostrar información de columnas
    print_log(f"* dfPrchss shape: {dfPrchss.shape}, columns: {list(dfPrchss.columns)}")
    print_log(f"* dfPrices shape: {dfPrices.shape}, columns: {list(dfPrices.columns)}")
    # Verificar columnas requeridas
    if not verify_columns(dfPrchss, ["Descripción"], "dfPrchss"):
        df["Marca"] = None
        df["Categoria"] = None
        return df
    if not verify_columns(dfPrices, ["Descripción", "Marca", "Categoria"], "dfPrices"):
        df["Marca"] = None
        df["Categoria"] = None
        return df
    try:
        # Realizar el merge
        print_log("df.merge()...")
        df = df.merge(
            dfPrices[["Descripción", "Marca", "Categoria"]], 
            on="Descripción", 
            how="left"
        )
        print_log(df[["Descripción", "Marca", "Categoria"]].head().to_string())
    except Exception as e:
        print_log(f"❌ Error durante el merge: {e}")
        # Fallback: añadir columnas vacías
        df["Marca"] = None
        df["Categoria"] = None
    return df

def procesar_precios(dfPrices, dfPrchss):
    """Procesa el dataframe de precios."""
    df = dfPrices.copy()
    print_log(f"df: {df}")
    df['P. Tienda'] = df['P. Tienda'].astype(float)
    df['C. Unit'] = df['C. Unit'].astype(float)
    df = df.merge(
        dfPrchss[["Descripción", "Cant", "% Desc", "Costo Final", "Pzs"]], 
        on="Descripción", 
        how="left"
    )
    print_log(f"df-updt: {df}")
    return df

def deep_clean_data(df):
    """Limpieza profunda de datos mejorada"""
    #print_log(f"* df: {df}")
    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]):
            # Conversión segura para columnas numéricas
            df[col] = df[col].apply(safe_convert_to_float)
        elif pd.api.types.is_datetime64_any_dtype(df[col]):
            # Manejo seguro de fechas
            df[col] = pd.to_datetime(df[col], errors='coerce').apply(
                lambda x: x.to_pydatetime() if pd.notna(x) else None
            )
    return df.replace([np.nan, pd.NA, 'None', 'none', 'NONE'], None)

def data_ingestion(dfCompras, dfPrecios):
    """Realiza la ingesta de datos a la base de datos."""
    conn = psycopg2.connect(**DB_CONFIG)
    success = True
    try:
        cur = conn.cursor()
        # Inicializar catálogos
        print_log("get_catalogs()...")
        get_catalogs(cur)
        previous_link = ""
        # Procesar cada compra
        for _, row in dfCompras.iterrows():
            #print_log(f"* row: {row}")
            # Obtener o crear tienda y proveedor
            str_link = row.get("Liga")
            print_log(f"str_link: {str_link}")
            if not str_link:
                print_log(f"previous_link: {previous_link}")
                str_link = previous_link
            previous_link = row.get("Liga")
            print_log("get_or_create_store()...")
            id_store = get_or_create_store(cur, str_link)
            if id_store is None:
                continue
            print_log(f"get_or_create_provider({id_store})...")
            id_provider = get_or_create_provider(cur, id_store, str_link)
            if id_provider is None:
                continue
            delivery_date = row.get("Fch Entrga")
            print_log(f"* delivery_date: {delivery_date}")
            if delivery_date is not None and "CANCELED" in str(delivery_date):
                continue
            # Obtener o crear producto
            product_name = row["Descripción"]
            print_log(f"* product_name: {product_name}")
            quantity = row["Cant"]
            unit_price = row["C. Unit"]
            if not product_name or not product_name:
                continue
            purchase_date = row["Fch Cmpr"]
            print_log(f"quantity: {quantity}, unit_price: {unit_price}, purchase_date: {purchase_date}")
            print_log(f"create_product({product_name})...")
            result = create_product(cur, row, product_name, "", quantity, unit_price, purchase_date)
            print_log(f"result: {result}")

            if not result["continue"]:
                continue

            id_product = result["id_product"]
            id_payment_type = get_id_payment_type("Tarjeta de Crédito")
            print_log(f"id_payment_type: {id_payment_type}")
            # Preparar datos de compra
            purchase_data = {
                "id_provider": id_provider,
                "id_payment_type": id_payment_type,
                "total": row["Total Cmpr"],
                "tax": 0,
                "ieps": 0,
                "purchase_date": row["Fch Cmpr"],
                "delivery_date": delivery_date,
                "exchange_rate": row.get("Dólar"),
                "shipping_cost": row.get("Envio", 0),
                "discount": row.get("Desct", 0)
            }
            # Insertar compra
            print_log("insert_purchase()...")
            id_purchase = insert_purchase(cur, purchase_data)
            # Preparar items de operación
            operation_items = [{
                "quantity": row["Cant"],
                "unit_price": row["C. Unit"],
                "unit_price_usd":  row.get("C. Unit US"),
                "discount_percentage": row.get("% Desc", 0),
                "pieces_per_unit": row.get("Pzs", 1),
                "final_cost": row.get("Costo Final"),
                "product_url": row.get("Liga", "")
            }]
            print_log(f"insert_operations({id_purchase})...")
            insert_operations(cur, id_purchase, id_product, operation_items)
            # Insertar precios si existe en el df de precios
            if row["Descripción"] in dfPrecios["Descripción"].values:
                price_row = dfPrecios[dfPrecios["Descripción"] == row["Descripción"]].iloc[0]
                operation = operation_items[0]
                price = price_row["P. Venta"]
                price = price if price else float(operation["final_cost"])*float(1+MARGEN_GANANCIA)
                offer_price = price_row.get("P. Oferta")
                offer_price = offer_price if offer_price else price*float(1-DESCUENTO_OFERTA)
                print_log(f"price: {price}, offer_price: {offer_price}")
                price_data = {
                    "price": price,
                    "offer_price": offer_price
                }
                print_log(f"insert_price({id_product})...")
                insert_price(cur, id_product, price_data)
        print("conn.commit()...")
        conn.commit()
        print("✅ Datos ingresados correctamente.")
    except Exception as e:
        print("conn.rollback()...")
        conn.rollback()
        print(f"❌ Error en la ingesta de datos: {e}")
        success = False
    finally:
        conn.close()
        return success

def procesar_archivo(filePath):
    """Procesa un archivo Excel y realiza la ingesta."""
    print(f"Procesando archivo: {filePath}")
    success = False
    xls = None
    wb = None
    try:
        print("extract_hyperlinks()...")
        # Extraer hipervínculos con manejo explícito del archivo
        with open(filePath, 'rb') as f:
            wb = load_workbook(f, data_only=True)
            links_urls = extract_hyperlinks(wb)  # Modificar extract_hyperlinks para aceptar wb
        # Leer datos con pandas asegurando cierre del archivo
        with pd.ExcelFile(filePath) as xls:
            df_prchss = pd.read_excel(xls, "Compras")
            df_prices = pd.read_excel(xls, "Precios")
        # Limpieza profunda
        print("deep_clean_data()...")
        df_prchss_cln = deep_clean_data(df_prchss)
        df_prices_cln = deep_clean_data(df_prices)
        print_log(f"df_prices_cln-len: {len(df_prices_cln)}")
        df_prchss_cln["Picture_URL"] = links_urls[:len(df_prchss_cln)]
        print_log(f"Picture_URL: {len(df_prchss_cln["Picture_URL"])}")
        print("procesar_purchase()...")
        df_prchss_upd = procesar_purchase(df_prchss_cln, df_prices_cln)
        print("procesar_precios()...")
        df_prices_upd = procesar_precios(df_prices_cln, df_prchss_cln)
        # Procesar ingesta
        success = data_ingestion(df_prchss_upd, df_prices_upd)
    except Exception as e:
        print(f"❌ Error procesando archivo {file_path}: {e}")
        success = False
    finally:
        # Cerrar explícitamente los recursos
        if wb is not None:
            wb.close()
        if xls is not None and hasattr(xls, 'close'):
            xls.close()
        # Mover el archivo solo después de cerrar todos los recursos
        move_file(filePath, success=success)
        return success

# ==== MAIN ====
if __name__ == "__main__":
    # Procesar todos los archivos XLSX en el directorio
    processed_count = 0
    error_count = 0
    for file_name in os.listdir(DATA_DIR):
        if file_name.endswith(".xlsx"):
            file_path = os.path.join(DATA_DIR, file_name)
            if procesar_archivo(file_path):
                processed_count += 1
            else:
                error_count += 1
    print(f"\n✅ Proceso de ingesta completado.")
    print(f"Archivos procesados correctamente: {processed_count}")
    print(f"Archivos con errores: {error_count}")
